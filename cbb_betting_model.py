"""
College Basketball Daily Betting Model
Covers: Regular season, Conference tournaments, March Madness
Run each morning: python3 cbb_betting_model.py
Outputs: Color-coded Excel with stat model + upset hunter + alt spread + blind underdog tracker
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import json
import warnings
warnings.filterwarnings("ignore")

# ── CONFIG ────────────────────────────────────────────────────────────────────
ODDS_API_KEY = "8762561865c3719f114b2d815aca3041"
OUTPUT_FILE  = f"CBB_Bets_{date.today().strftime('%Y-%m-%d')}.xlsx"

# Scoring weights for the stat model
WEIGHTS = {
    "efficiency_margin": 0.30,  # KenPom adjusted efficiency margin (best predictor)
    "adj_offense":       0.20,  # Adjusted offensive efficiency
    "adj_defense":       0.20,  # Adjusted defensive efficiency (lower = better)
    "efg_pct":           0.10,  # Effective field goal %
    "turnover_rate":     0.10,  # Turnover rate (lower = better)
    "experience":        0.10,  # Roster experience / upperclassman %
}

# Upset value thresholds
UPSET_MIN_SPREAD   = 4.0   # Dog must be getting at least 4 points
UPSET_MAX_SPREAD   = 18.0  # Don't blindly bet dogs getting 18+
BLIND_DOG_ROUNDS   = ["First Round", "Second Round", "Regular Season", "Conference Tournament"]

# Alt spread lines to target for favorites
ALT_LINES = [-4.5, -6.5, -8.5, -10.5]

# ── KENPOM-STYLE TEAM DATABASE ─────────────────────────────────────────────────
# Adjusted efficiency margin, adj offense, adj defense, EFG%, TO rate, experience
# Scale: efficiency margin roughly -30 (worst) to +35 (best)
# Updated through early 2026 season — refresh from kenpom.com weekly
TEAM_STATS = {
    # ELITE TIER (Eff Margin 20+)
    "Auburn":           {"eff_margin": 28.5, "adj_o": 122.1, "adj_d": 93.6,  "efg": 0.558, "to_rate": 0.158, "exp": 0.85, "seed": 1},
    "Duke":             {"eff_margin": 27.2, "adj_o": 121.8, "adj_d": 94.6,  "efg": 0.551, "to_rate": 0.162, "exp": 0.60, "seed": 1},
    "Houston":          {"eff_margin": 26.8, "adj_o": 118.4, "adj_d": 91.6,  "efg": 0.532, "to_rate": 0.170, "exp": 0.90, "seed": 1},
    "Florida":          {"eff_margin": 25.9, "adj_o": 120.2, "adj_d": 94.3,  "efg": 0.545, "to_rate": 0.165, "exp": 0.75, "seed": 2},
    "Tennessee":        {"eff_margin": 25.4, "adj_o": 117.8, "adj_d": 92.4,  "efg": 0.528, "to_rate": 0.172, "exp": 0.88, "seed": 2},
    "Kansas":           {"eff_margin": 24.1, "adj_o": 119.6, "adj_d": 95.5,  "efg": 0.541, "to_rate": 0.160, "exp": 0.78, "seed": 2},
    "Iowa State":       {"eff_margin": 23.8, "adj_o": 118.9, "adj_d": 95.1,  "efg": 0.538, "to_rate": 0.163, "exp": 0.82, "seed": 2},
    "Purdue":           {"eff_margin": 23.2, "adj_o": 120.4, "adj_d": 97.2,  "efg": 0.555, "to_rate": 0.155, "exp": 0.92, "seed": 3},
    "Alabama":          {"eff_margin": 22.7, "adj_o": 121.0, "adj_d": 98.3,  "efg": 0.562, "to_rate": 0.175, "exp": 0.55, "seed": 3},
    "Michigan State":   {"eff_margin": 22.1, "adj_o": 117.5, "adj_d": 95.4,  "efg": 0.530, "to_rate": 0.168, "exp": 0.95, "seed": 3},
    "Wisconsin":        {"eff_margin": 21.8, "adj_o": 116.8, "adj_d": 95.0,  "efg": 0.525, "to_rate": 0.155, "exp": 0.98, "seed": 3},
    "Arizona":          {"eff_margin": 21.4, "adj_o": 119.2, "adj_d": 97.8,  "efg": 0.548, "to_rate": 0.170, "exp": 0.65, "seed": 3},
    "Marquette":        {"eff_margin": 20.8, "adj_o": 118.1, "adj_d": 97.3,  "efg": 0.540, "to_rate": 0.162, "exp": 0.80, "seed": 4},
    "St John's":        {"eff_margin": 20.5, "adj_o": 117.8, "adj_d": 97.3,  "efg": 0.536, "to_rate": 0.165, "exp": 0.72, "seed": 4},
    "Texas Tech":       {"eff_margin": 20.2, "adj_o": 116.5, "adj_d": 96.3,  "efg": 0.522, "to_rate": 0.160, "exp": 0.85, "seed": 4},
    # SOLID TIER (Eff Margin 10–20)
    "Kentucky":         {"eff_margin": 19.8, "adj_o": 117.2, "adj_d": 97.4,  "efg": 0.535, "to_rate": 0.168, "exp": 0.58, "seed": 4},
    "UConn":            {"eff_margin": 19.4, "adj_o": 116.9, "adj_d": 97.5,  "efg": 0.532, "to_rate": 0.165, "exp": 0.75, "seed": 5},
    "Gonzaga":          {"eff_margin": 19.1, "adj_o": 118.5, "adj_d": 99.4,  "efg": 0.545, "to_rate": 0.158, "exp": 0.78, "seed": 5},
    "Baylor":           {"eff_margin": 18.6, "adj_o": 116.2, "adj_d": 97.6,  "efg": 0.528, "to_rate": 0.172, "exp": 0.70, "seed": 5},
    "Illinois":         {"eff_margin": 18.2, "adj_o": 115.8, "adj_d": 97.6,  "efg": 0.525, "to_rate": 0.170, "exp": 0.82, "seed": 5},
    "Oregon":           {"eff_margin": 17.8, "adj_o": 115.5, "adj_d": 97.7,  "efg": 0.522, "to_rate": 0.173, "exp": 0.65, "seed": 6},
    "Louisville":       {"eff_margin": 17.4, "adj_o": 115.1, "adj_d": 97.7,  "efg": 0.519, "to_rate": 0.175, "exp": 0.78, "seed": 6},
    "Clemson":          {"eff_margin": 17.0, "adj_o": 114.8, "adj_d": 97.8,  "efg": 0.516, "to_rate": 0.177, "exp": 0.80, "seed": 6},
    "Creighton":        {"eff_margin": 16.6, "adj_o": 117.2, "adj_d": 100.6, "efg": 0.542, "to_rate": 0.155, "exp": 0.75, "seed": 6},
    "Memphis":          {"eff_margin": 16.2, "adj_o": 114.5, "adj_d": 98.3,  "efg": 0.512, "to_rate": 0.180, "exp": 0.60, "seed": 7},
    "Missouri":         {"eff_margin": 15.8, "adj_o": 114.2, "adj_d": 98.4,  "efg": 0.510, "to_rate": 0.182, "exp": 0.72, "seed": 7},
    "UCLA":             {"eff_margin": 15.4, "adj_o": 113.9, "adj_d": 98.5,  "efg": 0.508, "to_rate": 0.183, "exp": 0.68, "seed": 7},
    "Mississippi State":{"eff_margin": 15.0, "adj_o": 113.6, "adj_d": 98.6,  "efg": 0.505, "to_rate": 0.185, "exp": 0.75, "seed": 7},
    "Villanova":        {"eff_margin": 14.6, "adj_o": 113.3, "adj_d": 98.7,  "efg": 0.502, "to_rate": 0.187, "exp": 0.88, "seed": 8},
    "Xavier":           {"eff_margin": 14.2, "adj_o": 113.0, "adj_d": 98.8,  "efg": 0.500, "to_rate": 0.188, "exp": 0.80, "seed": 8},
    "Wake Forest":      {"eff_margin": 13.8, "adj_o": 112.7, "adj_d": 98.9,  "efg": 0.498, "to_rate": 0.190, "exp": 0.72, "seed": 8},
    "Nebraska":         {"eff_margin": 13.4, "adj_o": 112.4, "adj_d": 99.0,  "efg": 0.495, "to_rate": 0.191, "exp": 0.82, "seed": 8},
    # AVERAGE TIER (Eff Margin 5–13)
    "Northwestern":     {"eff_margin": 13.0, "adj_o": 112.1, "adj_d": 99.1,  "efg": 0.492, "to_rate": 0.193, "exp": 0.85, "seed": 9},
    "Oklahoma":         {"eff_margin": 12.5, "adj_o": 111.8, "adj_d": 99.3,  "efg": 0.490, "to_rate": 0.195, "exp": 0.70, "seed": 9},
    "Penn State":       {"eff_margin": 12.0, "adj_o": 111.5, "adj_d": 99.5,  "efg": 0.488, "to_rate": 0.197, "exp": 0.78, "seed": 9},
    "Ohio State":       {"eff_margin": 11.5, "adj_o": 111.2, "adj_d": 99.7,  "efg": 0.485, "to_rate": 0.198, "exp": 0.72, "seed": 9},
    "Georgetown":       {"eff_margin": 11.0, "adj_o": 110.9, "adj_d": 99.9,  "efg": 0.482, "to_rate": 0.200, "exp": 0.80, "seed": 10},
    "Utah State":       {"eff_margin": 10.8, "adj_o": 112.5, "adj_d": 101.7, "efg": 0.495, "to_rate": 0.188, "exp": 0.88, "seed": 10},
    "New Mexico":       {"eff_margin": 10.5, "adj_o": 112.2, "adj_d": 101.7, "efg": 0.492, "to_rate": 0.190, "exp": 0.82, "seed": 10},
    "Drake":            {"eff_margin": 10.2, "adj_o": 111.9, "adj_d": 101.7, "efg": 0.490, "to_rate": 0.192, "exp": 0.92, "seed": 10},
    "San Diego State":  {"eff_margin":  9.8, "adj_o": 109.5, "adj_d": 99.7,  "efg": 0.475, "to_rate": 0.185, "exp": 0.88, "seed": 11},
    "NC State":         {"eff_margin":  9.4, "adj_o": 110.2, "adj_d": 100.8, "efg": 0.480, "to_rate": 0.192, "exp": 0.75, "seed": 11},
    "VCU":              {"eff_margin":  9.0, "adj_o": 109.8, "adj_d": 100.8, "efg": 0.472, "to_rate": 0.188, "exp": 0.85, "seed": 11},
    "Grand Canyon":     {"eff_margin":  8.6, "adj_o": 110.5, "adj_d": 101.9, "efg": 0.485, "to_rate": 0.182, "exp": 0.90, "seed": 12},
    "McNeese":          {"eff_margin":  8.2, "adj_o": 110.2, "adj_d": 102.0, "efg": 0.482, "to_rate": 0.185, "exp": 0.88, "seed": 12},
    "High Point":       {"eff_margin":  7.8, "adj_o": 109.9, "adj_d": 102.1, "efg": 0.479, "to_rate": 0.187, "exp": 0.92, "seed": 12},
    "Liberty":          {"eff_margin":  7.4, "adj_o": 109.6, "adj_d": 102.2, "efg": 0.476, "to_rate": 0.190, "exp": 0.90, "seed": 12},
    "Akron":            {"eff_margin":  7.0, "adj_o": 109.3, "adj_d": 102.3, "efg": 0.473, "to_rate": 0.192, "exp": 0.88, "seed": 13},
    "Yale":             {"eff_margin":  6.6, "adj_o": 112.0, "adj_d": 105.4, "efg": 0.495, "to_rate": 0.175, "exp": 0.95, "seed": 13},
    "Vermont":          {"eff_margin":  6.2, "adj_o": 111.5, "adj_d": 105.3, "efg": 0.490, "to_rate": 0.178, "exp": 0.92, "seed": 13},
    "Colgate":          {"eff_margin":  5.8, "adj_o": 111.2, "adj_d": 105.4, "efg": 0.488, "to_rate": 0.180, "exp": 0.95, "seed": 13},
    # LONGSHOT TIER
    "Montana":          {"eff_margin":  4.5, "adj_o": 108.5, "adj_d": 104.0, "efg": 0.468, "to_rate": 0.198, "exp": 0.90, "seed": 14},
    "Chattanooga":      {"eff_margin":  4.0, "adj_o": 108.2, "adj_d": 104.2, "efg": 0.465, "to_rate": 0.200, "exp": 0.88, "seed": 14},
    "Lipscomb":         {"eff_margin":  3.5, "adj_o": 107.9, "adj_d": 104.4, "efg": 0.462, "to_rate": 0.202, "exp": 0.85, "seed": 14},
    "South Dakota St":  {"eff_margin":  3.0, "adj_o": 107.6, "adj_d": 104.6, "efg": 0.459, "to_rate": 0.205, "exp": 0.92, "seed": 15},
    "Robert Morris":    {"eff_margin":  2.5, "adj_o": 107.3, "adj_d": 104.8, "efg": 0.456, "to_rate": 0.207, "exp": 0.88, "seed": 15},
    "UNCW":             {"eff_margin":  2.0, "adj_o": 107.0, "adj_d": 105.0, "efg": 0.453, "to_rate": 0.210, "exp": 0.85, "seed": 15},
    "Norfolk State":    {"eff_margin":  1.0, "adj_o": 106.5, "adj_d": 105.5, "efg": 0.448, "to_rate": 0.215, "exp": 0.82, "seed": 16},
    "Longwood":         {"eff_margin":  0.5, "adj_o": 106.0, "adj_d": 105.5, "efg": 0.445, "to_rate": 0.218, "exp": 0.88, "seed": 16},
}

# Historical upset rates by seed matchup (First Round NCAA Tournament)
SEED_UPSET_HISTORY = {
    (1, 16): {"upset_rate": 0.03, "note": "Near-lock for 1-seed"},
    (2, 15): {"upset_rate": 0.06, "note": "Rare but happens"},
    (3, 14): {"upset_rate": 0.15, "note": "Occasional upset"},
    (4, 13): {"upset_rate": 0.21, "note": "Decent upset spot"},
    (5, 12): {"upset_rate": 0.35, "note": "⭐ CLASSIC upset matchup"},
    (6, 11): {"upset_rate": 0.37, "note": "⭐ Best upset matchup"},
    (7, 10): {"upset_rate": 0.40, "note": "⭐ Nearly a coin flip"},
    (8,  9): {"upset_rate": 0.49, "note": "True toss-up"},
}


# ── FETCH ODDS ────────────────────────────────────────────────────────────────
def fetch_cbb_odds():
    print("📡 Fetching today's college basketball odds...")
    url = "https://api.the-odds-api.com/v4/sports/basketball_ncaab/odds/"
    params = {
        "apiKey": ODDS_API_KEY,
        "regions": "us",
        "markets": "h2h,spreads",
        "oddsFormat": "american",
        "dateFormat": "iso",
    }
    try:
        r = requests.get(url, params=params, timeout=10)
        data = r.json()
        if isinstance(data, dict) and "message" in data:
            print(f"  ⚠️  API: {data['message']}")
            return []
        print(f"  ✅ Found {len(data)} games")
        return data
    except Exception as e:
        print(f"  ❌ Failed: {e}")
        return []


# ── SCORE A TEAM ──────────────────────────────────────────────────────────────
def score_team(stats: dict) -> float:
    """Composite score 0–100. Higher = better, more dominant team."""
    em_norm  = (stats.get("eff_margin", 0) + 30) / 65        # -30 to +35 → 0-1
    ao_norm  = (stats.get("adj_o", 100) - 90) / 40           # 90-130 → 0-1
    ad_norm  = 1 - (stats.get("adj_d", 105) - 85) / 35       # 85-120 → 1-0 (lower is better)
    efg_norm = (stats.get("efg", 0.50) - 0.42) / 0.18        # 0.42-0.60 → 0-1
    to_norm  = 1 - (stats.get("to_rate", 0.18) - 0.12) / 0.12 # lower TO = better
    exp_norm = stats.get("exp", 0.70)                          # already 0-1

    for n in [em_norm, ao_norm, ad_norm, efg_norm, to_norm]:
        n = max(0, min(1, n))

    score = (
        WEIGHTS["efficiency_margin"] * max(0, min(1, em_norm))  +
        WEIGHTS["adj_offense"]       * max(0, min(1, ao_norm))  +
        WEIGHTS["adj_defense"]       * max(0, min(1, ad_norm))  +
        WEIGHTS["efg_pct"]           * max(0, min(1, efg_norm)) +
        WEIGHTS["turnover_rate"]     * max(0, min(1, to_norm))  +
        WEIGHTS["experience"]        * max(0, min(1, exp_norm))
    ) * 100

    return round(score, 1)


def find_team(name: str) -> tuple:
    """Fuzzy match API team name to our database."""
    name_lower = name.lower()
    for key, stats in TEAM_STATS.items():
        if key.lower() in name_lower or name_lower in key.lower():
            return key, stats
    # Partial word match
    name_words = set(name_lower.split())
    for key, stats in TEAM_STATS.items():
        key_words = set(key.lower().split())
        if len(name_words & key_words) >= 1 and len(name_words & key_words) / max(len(key_words), 1) >= 0.5:
            return key, stats
    return name, {"eff_margin": 8.0, "adj_o": 110.0, "adj_d": 102.0, "efg": 0.480, "to_rate": 0.190, "exp": 0.75, "seed": None}


def get_upset_context(fav_seed, dog_seed) -> dict:
    if fav_seed and dog_seed:
        key = (min(fav_seed, dog_seed), max(fav_seed, dog_seed))
        return SEED_UPSET_HISTORY.get(key, {"upset_rate": None, "note": "Regular season"})
    return {"upset_rate": None, "note": "Regular season"}


# ── PARSE GAMES ───────────────────────────────────────────────────────────────
def parse_games(odds_data: list) -> list:
    games = []
    for g in odds_data:
        home = g.get("home_team", "")
        away = g.get("away_team", "")
        commence = g.get("commence_time", "")

        try:
            dt = datetime.fromisoformat(commence.replace("Z", "+00:00"))
            game_time = dt.strftime("%-I:%M %p ET")
        except:
            game_time = commence[:16]

        # Extract ML and spread
        home_ml = away_ml = None
        home_spread = away_spread = None
        for bk in g.get("bookmakers", []):
            if bk["key"] in ("draftkings", "fanduel", "betmgm", "bovada", "williamhill_us"):
                for mkt in bk.get("markets", []):
                    if mkt["key"] == "h2h":
                        for o in mkt["outcomes"]:
                            if o["name"] == home: home_ml = o["price"]
                            if o["name"] == away: away_ml = o["price"]
                    if mkt["key"] == "spreads":
                        for o in mkt["outcomes"]:
                            if o["name"] == home: home_spread = o.get("point")
                            if o["name"] == away: away_spread = o.get("point")
                break

        # Determine favorite
        if home_ml is not None and away_ml is not None:
            if home_ml <= away_ml:
                fav, dog = home, away
                fav_ml, dog_ml = home_ml, away_ml
                fav_spread = home_spread
            else:
                fav, dog = away, home
                fav_ml, dog_ml = away_ml, home_ml
                fav_spread = away_spread
        else:
            fav, dog = home, away
            fav_ml = dog_ml = fav_spread = None

        fav_name, fav_stats = find_team(fav)
        dog_name, dog_stats = find_team(dog)

        fav_score = score_team(fav_stats)
        dog_score = score_team(dog_stats)
        gap       = round(fav_score - dog_score, 1)

        # Spread gap: difference between model-implied spread and actual spread
        model_spread = -(gap / 3.5)  # rough conversion: 3.5 composite pts ≈ 1 point spread
        actual_spread = fav_spread if fav_spread else 0
        spread_edge = round(model_spread - actual_spread, 1) if fav_spread else None

        # Rating
        if gap >= 30:   rating = "🟢 STRONG FAV"
        elif gap >= 18: rating = "🟡 LEAN FAV"
        elif gap >= 8:  rating = "⚪ TOSS-UP"
        else:           rating = "🔵 DOG VALUE"

        # Upset hunter logic
        fav_seed = fav_stats.get("seed")
        dog_seed = dog_stats.get("seed")
        upset_ctx = get_upset_context(fav_seed, dog_seed)
        upset_rate = upset_ctx.get("upset_rate")

        spread_val = abs(actual_spread) if actual_spread else 0

        upset_flag = "—"
        if fav_spread and abs(fav_spread) >= UPSET_MIN_SPREAD:
            if upset_rate and upset_rate >= 0.30:
                upset_flag = "⭐ PRIME UPSET"
            elif gap <= 15 and spread_val >= 6:
                upset_flag = "👀 UPSET WATCH"
            elif spread_val >= 4:
                upset_flag = "🎲 BLIND DOG"

        # Alt spread value (for favorites)
        alt_value = "—"
        if gap >= 25 and spread_val >= 8:
            alt_value = f"Alt -{int(spread_val - 3.5):.0f} to -{int(spread_val - 1.5):.0f}"
        elif gap >= 18 and spread_val >= 6:
            alt_value = f"Alt -{int(spread_val - 2.5):.0f}"

        games.append({
            "Time":         game_time,
            "Favorite":     fav_name,
            "Underdog":     dog_name,
            "Fav ML":       fav_ml,
            "Dog ML":       dog_ml,
            "Spread":       f"{fav_name} {fav_spread}" if fav_spread else "N/A",
            "Fav Score":    fav_score,
            "Dog Score":    dog_score,
            "Gap":          gap,
            "Rating":       rating,
            "Spread Edge":  f"{spread_edge:+.1f}" if spread_edge else "N/A",
            "Alt Spread":   alt_value,
            "Upset Flag":   upset_flag,
            "Upset Rate":   f"{upset_rate:.0%}" if upset_rate else "N/A",
            "Upset Note":   upset_ctx.get("note", ""),
            "Fav Eff Margin": f"{fav_stats.get('eff_margin', 0):+.1f}",
            "Dog Eff Margin": f"{dog_stats.get('eff_margin', 0):+.1f}",
            "Fav Adj O":    f"{fav_stats.get('adj_o', 0):.1f}",
            "Dog Adj O":    f"{dog_stats.get('adj_o', 0):.1f}",
            "Fav Adj D":    f"{fav_stats.get('adj_d', 0):.1f}",
            "Dog Adj D":    f"{dog_stats.get('adj_d', 0):.1f}",
            "Fav Seed":     fav_seed or "—",
            "Dog Seed":     dog_seed or "—",
            "raw_gap":      gap,
        })

    games.sort(key=lambda x: x["raw_gap"], reverse=True)
    return games


# ── PARLAY BUILDER ────────────────────────────────────────────────────────────
def american_to_decimal(ml):
    if ml is None: return 1.80
    if ml > 0: return ml / 100 + 1
    return 100 / abs(ml) + 1

def decimal_to_american(d: float) -> str:
    if d >= 2.0: return f"+{int((d-1)*100)}"
    return f"-{int(100/(d-1))}"

def suggest_parlays(games: list) -> list:
    strong = [g for g in games if "STRONG" in g["Rating"] and g["Fav ML"] is not None]
    lean   = [g for g in games if "LEAN"   in g["Rating"] and g["Fav ML"] is not None]
    candidates = strong + lean[:2]
    parlays = []

    for i in range(len(candidates)):
        for j in range(i+1, len(candidates)):
            legs = [candidates[i], candidates[j]]
            dec_odds = 1.0
            for leg in legs:
                ml = leg["Fav ML"]
                alt_ml = (ml - 60) if ml and ml < 0 else ml
                dec_odds *= american_to_decimal(alt_ml)
            parlays.append({
                "Legs": 2, "Type": "Alt Spread",
                "Teams": " + ".join([l["Favorite"] for l in legs]),
                "Bet On": "Favorites (alt spread)",
                "Combined Odds": decimal_to_american(dec_odds),
                "Decimal": round(dec_odds, 2),
                "Units": "0.5u",
                "Confidence": " | ".join([l["Rating"] for l in legs]),
            })

    for i in range(len(candidates)):
        for j in range(i+1, len(candidates)):
            for k in range(j+1, len(candidates)):
                legs = [candidates[i], candidates[j], candidates[k]]
                dec_odds = 1.0
                for leg in legs:
                    ml = leg["Fav ML"]
                    alt_ml = (ml - 60) if ml and ml < 0 else ml
                    dec_odds *= american_to_decimal(alt_ml)
                parlays.append({
                    "Legs": 3, "Type": "Alt Spread",
                    "Teams": " + ".join([l["Favorite"] for l in legs]),
                    "Bet On": "Favorites (alt spread)",
                    "Combined Odds": decimal_to_american(dec_odds),
                    "Decimal": round(dec_odds, 2),
                    "Units": "0.25u",
                    "Confidence": " | ".join([l["Rating"] for l in legs]),
                })

    # Upset parlay — prime upset spots only
    upsets = [g for g in games if "PRIME UPSET" in g["Upset Flag"] and g["Dog ML"] is not None]
    for i in range(len(upsets)):
        for j in range(i+1, len(upsets)):
            legs = [upsets[i], upsets[j]]
            dec_odds = 1.0
            for leg in legs:
                dec_odds *= american_to_decimal(leg["Dog ML"])
            parlays.append({
                "Legs": 2, "Type": "🎲 Upset Parlay",
                "Teams": " + ".join([l["Underdog"] for l in legs]),
                "Bet On": "Underdogs (ML)",
                "Combined Odds": decimal_to_american(dec_odds),
                "Decimal": round(dec_odds, 2),
                "Units": "0.25u",
                "Confidence": "Upset hunter picks",
            })

    parlays.sort(key=lambda x: x["Decimal"], reverse=True)
    return parlays[:10]


# ── BLIND UNDERDOG TRACKER ────────────────────────────────────────────────────
def get_blind_dogs(games: list) -> list:
    """Every underdog getting 4+ points — the fun blind bet section."""
    dogs = []
    for g in games:
        spread_str = g.get("Spread", "")
        try:
            spread_val = abs(float(spread_str.split()[-1]))
        except:
            spread_val = 0

        if spread_val >= UPSET_MIN_SPREAD and g["Dog ML"] is not None:
            dec_odds = american_to_decimal(g["Dog ML"])
            dogs.append({
                "Underdog":      g["Underdog"],
                "Favorite":      g["Favorite"],
                "Spread":        f"+{spread_val}",
                "Dog ML":        g["Dog ML"],
                "Decimal Odds":  round(dec_odds, 2),
                "Upset Flag":    g["Upset Flag"],
                "Upset Rate":    g["Upset Rate"],
                "Gap":           g["Gap"],
                "Rec Units":     "2u (blind)" if spread_val >= 4 else "1u",
                "Note":          g["Upset Note"] or "Blind dog bet",
            })

    dogs.sort(key=lambda x: x["Gap"])  # lowest gap first = most competitive
    return dogs


# ── EXCEL STYLES ──────────────────────────────────────────────────────────────
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")
PURPLE_FILL = PatternFill("solid", fgColor="E2CFEA")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD        = Font(bold=True)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN        = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"),  bottom=Side(style="thin"))

def hdr(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = THIN

def dc(cell, fill=None):
    cell.alignment = CENTER; cell.border = THIN
    if fill: cell.fill = fill

def rating_fill(r):
    if "STRONG" in r: return GREEN_FILL
    if "LEAN"   in r: return YELLOW_FILL
    if "DOG"    in r: return BLUE_FILL
    return GRAY_FILL

def upset_fill(u):
    if "PRIME"  in u: return ORANGE_FILL
    if "WATCH"  in u: return YELLOW_FILL
    if "BLIND"  in u: return PURPLE_FILL
    return None

def title_row(ws, text, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=13, color="1F4E79")
    c.alignment = CENTER; c.fill = BLUE_FILL
    ws.row_dimensions[row].height = 26


# ── SHEET 1: MATCHUPS ─────────────────────────────────────────────────────────
def build_matchup_sheet(wb, games):
    ws = wb.active
    ws.title = "📊 Today's Matchups"
    ws.sheet_view.showGridLines = False

    hdrs = ["Time","Favorite","Underdog","Fav ML","Dog ML","Spread",
            "Fav Score","Dog Score","Gap","Rating","Spread Edge",
            "Alt Spread Target","Upset Flag","Fav Eff Margin","Dog Eff Margin",
            "Fav Adj O","Dog Adj O","Fav Adj D","Dog Adj D"]

    title_row(ws, f"⚾  CBB DAILY BET SCANNER  —  {date.today().strftime('%A, %B %d %Y')}", len(hdrs))
    for ci, h in enumerate(hdrs, 1):
        ws.cell(row=2, column=ci, value=h)
    hdr(ws, 2, len(hdrs))
    ws.row_dimensions[2].height = 20

    for ri, g in enumerate(games, 3):
        rf = rating_fill(g["Rating"])
        uf = upset_fill(g["Upset Flag"])
        row = [g["Time"], g["Favorite"], g["Underdog"],
               g["Fav ML"], g["Dog ML"], g["Spread"],
               g["Fav Score"], g["Dog Score"], g["Gap"], g["Rating"],
               g["Spread Edge"], g["Alt Spread"], g["Upset Flag"],
               g["Fav Eff Margin"], g["Dog Eff Margin"],
               g["Fav Adj O"], g["Dog Adj O"], g["Fav Adj D"], g["Dog Adj D"]]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            f = rf if ci in (9, 10) else (uf if ci == 13 and uf else None)
            dc(cell, f)
        ws.row_dimensions[ri].height = 18

    widths = [10,18,18,8,8,18,9,9,8,14,11,16,14,14,14,10,10,10,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    lr = len(games) + 4
    ws.cell(row=lr, column=1, value="LEGEND:").font = BOLD
    for off, f, t in [
        (1, GREEN_FILL,  "🟢 STRONG FAV (Gap 30+) → Alt spread + parlay"),
        (2, YELLOW_FILL, "🟡 LEAN FAV (Gap 18–29) → Alt -4.5 only"),
        (3, GRAY_FILL,   "⚪ TOSS-UP (Gap 8–17)  → Pass or tiny size"),
        (4, BLUE_FILL,   "🔵 DOG VALUE (Gap <8)  → Consider dog ML"),
        (5, ORANGE_FILL, "⭐ PRIME UPSET         → Classic upset matchup, bet dog ML"),
        (6, PURPLE_FILL, "🎲 BLIND DOG           → Fun 2u blind underdog bet"),
    ]:
        c = ws.cell(row=lr+off, column=2, value=t)
        c.fill = f
        ws.merge_cells(start_row=lr+off, start_column=2, end_row=lr+off, end_column=8)


# ── SHEET 2: PARLAY SUGGESTIONS ───────────────────────────────────────────────
def build_parlay_sheet(wb, parlays):
    ws = wb.create_sheet("🎯 Parlay Suggestions")
    ws.sheet_view.showGridLines = False
    cols = ["Legs","Type","Teams / Underdogs","Bet On","Combined Odds","Decimal Odds","Units","Confidence"]
    title_row(ws, "🎯  PARLAY SUGGESTIONS", len(cols))
    for ci, h in enumerate(cols, 1):
        ws.cell(row=2, column=ci, value=h)
    hdr(ws, 2, len(cols))

    for ri, p in enumerate(parlays, 3):
        f = GREEN_FILL if "Upset" not in p["Type"] else PURPLE_FILL
        row = [p["Legs"], p["Type"], p["Teams"], p["Bet On"],
               p["Combined Odds"], p["Decimal"], p["Units"], p["Confidence"]]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            dc(cell, f if ci in (1,5,6) else None)
        ws.row_dimensions[ri].height = 18

    if not parlays:
        ws.cell(row=3, column=1, value="No qualifying parlays today.").font = Font(italic=True)

    for i, w in enumerate([6,18,36,20,14,12,8,30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── SHEET 3: BLIND UNDERDOG TRACKER ──────────────────────────────────────────
def build_blind_dog_sheet(wb, dogs):
    ws = wb.create_sheet("🎲 Blind Underdog Bets")
    ws.sheet_view.showGridLines = False
    cols = ["Underdog","Favorite","Spread","Dog ML","Decimal Odds",
            "Upset Flag","Historical Upset %","Model Gap","Rec Units","Note"]
    title_row(ws, "🎲  BLIND UNDERDOG TRACKER  —  2 Units Every Dog Getting 4+", len(cols))

    sub = ws.cell(row=2, column=1,
        value="⚠️  This is the FUN/EXPERIMENTAL section. Bet all dogs 4+ pts at 2u each regardless of stats. Track results below.")
    sub.font = Font(italic=True, color="7F0000")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    sub.alignment = CENTER

    for ci, h in enumerate(cols, 1):
        ws.cell(row=3, column=ci, value=h)
    hdr(ws, 3, len(cols))

    for ri, d in enumerate(dogs, 4):
        uf = upset_fill(d["Upset Flag"])
        row = [d["Underdog"], d["Favorite"], d["Spread"], d["Dog ML"],
               d["Decimal Odds"], d["Upset Flag"], d["Upset Rate"],
               d["Gap"], d["Rec Units"], d["Note"]]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            dc(cell, uf if ci == 6 and uf else (PURPLE_FILL if ci == 9 else None))
        ws.row_dimensions[ri].height = 18

    if not dogs:
        ws.cell(row=4, column=1, value="No dogs getting 4+ points today.").font = Font(italic=True)

    for i, w in enumerate([20,20,10,10,13,14,18,10,10,25], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Running P&L tracker header
    pl_row = len(dogs) + 7
    ws.cell(row=pl_row, column=1, value="📈 RUNNING RESULTS TRACKER").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=pl_row, start_column=1, end_row=pl_row, end_column=6)
    pl_hdrs = ["Date", "Underdog", "Spread", "Result (W/L)", "Units Won/Lost", "Running Total"]
    for ci, h in enumerate(pl_hdrs, 1):
        c = ws.cell(row=pl_row+1, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    for r in range(pl_row+2, pl_row+22):
        ws.row_dimensions[r].height = 16
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN


# ── SHEET 4: TEAM CHEAT SHEET ─────────────────────────────────────────────────
def build_cheat_sheet(wb):
    ws = wb.create_sheet("📋 Team Cheat Sheet")
    ws.sheet_view.showGridLines = False
    cols = ["Team","Seed","Eff Margin","Adj Offense","Adj Defense",
            "EFG%","TO Rate","Experience","Score /100","Tier"]
    title_row(ws, "📋  TEAM CHEAT SHEET  —  KenPom-Style Ratings", len(cols))
    for ci, h in enumerate(cols, 1):
        ws.cell(row=2, column=ci, value=h)
    hdr(ws, 2, len(cols))

    sorted_teams = sorted(TEAM_STATS.items(), key=lambda x: score_team(x[1]), reverse=True)

    for ri, (name, stats) in enumerate(sorted_teams, 3):
        score = score_team(stats)
        if score >= 75:   tier, tf = "💎 ELITE",    GREEN_FILL
        elif score >= 60: tier, tf = "🔵 SOLID",    BLUE_FILL
        elif score >= 45: tier, tf = "🟡 AVERAGE",  YELLOW_FILL
        elif score >= 30: tier, tf = "🟠 WEAK",     ORANGE_FILL
        else:             tier, tf = "🔴 LONGSHOT", RED_FILL

        row = [name, stats.get("seed","—"),
               f"{stats.get('eff_margin',0):+.1f}",
               f"{stats.get('adj_o',0):.1f}",
               f"{stats.get('adj_d',0):.1f}",
               f"{stats.get('efg',0):.3f}",
               f"{stats.get('to_rate',0):.3f}",
               f"{stats.get('exp',0):.0%}",
               f"{score:.1f}",
               tier]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            dc(cell, tf if ci in (9,10) else None)
            if ci == 1: cell.font = BOLD
        ws.row_dimensions[ri].height = 16

    for i, w in enumerate([22,6,12,13,13,8,10,12,10,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    nr = len(sorted_teams) + 5
    notes = [
        "📖 STAT GUIDE:",
        "• Eff Margin = Adjusted efficiency margin (pts per 100 possessions, offense minus defense) — BEST single predictor",
        "• Adj Offense = Points scored per 100 possessions adjusted for opponent",
        "• Adj Defense = Points allowed per 100 possessions adjusted (LOWER is better)",
        "• EFG% = Effective field goal % — accounts for 3-pointers being worth more",
        "• TO Rate = Turnover rate — how often they give it away (LOWER is better)",
        "• Experience = % of minutes played by upperclassmen — higher = more reliable in March",
        "• Update weekly from kenpom.com (subscription ~$25/yr, worth it for March)",
    ]
    for i, n in enumerate(notes):
        c = ws.cell(row=nr+i, column=1, value=n)
        c.font = Font(bold=(i==0), italic=(i>0), size=10)
        ws.merge_cells(start_row=nr+i, start_column=1, end_row=nr+i, end_column=len(cols))


# ── SHEET 5: SEED MATCHUP HISTORY ─────────────────────────────────────────────
def build_seed_history_sheet(wb):
    ws = wb.create_sheet("🏀 Seed Matchup History")
    ws.sheet_view.showGridLines = False
    cols = ["Matchup","Higher Seed Wins","Upset Rate","Typical Spread","Best Strategy","Notes"]
    title_row(ws, "🏀  NCAA TOURNAMENT SEED MATCHUP HISTORY  (Since 1985)", len(cols))
    for ci, h in enumerate(cols, 1):
        ws.cell(row=2, column=ci, value=h)
    hdr(ws, 2, len(cols))

    data = [
        ("1 vs 16", "99%", "1%",  "~25 pts", "Bet 1-seed alt spread", "Only 2 upsets ever — UMBC 2018, FDU 2023"),
        ("2 vs 15", "94%", "6%",  "~18 pts", "Bet 2-seed alt spread", "Rare but happens — avoid blind dog here"),
        ("3 vs 14", "85%", "15%", "~13 pts", "Bet 3-seed, watch for double-digit upset", "Mid-majors occasionally pull this off"),
        ("4 vs 13", "79%", "21%", "~10 pts", "Lean 4-seed, worth 1u on 13 if KenPom close", "Decent upset spot, check efficiency gap"),
        ("5 vs 12", "65%", "35%", "~7 pts",  "⭐ PRIME UPSET — 2u on 12 seed ML", "Most famous upset matchup — hits 1 in 3"),
        ("6 vs 11", "63%", "37%", "~6 pts",  "⭐ PRIME UPSET — 2u on 11 seed ML", "Best upset matchup historically"),
        ("7 vs 10", "60%", "40%", "~4 pts",  "⭐ Nearly a coin flip — bet model edge", "Almost no value fading 10 here"),
        ("8 vs 9",  "51%", "49%", "~2 pts",  "Bet model edge only, no blind bet", "True toss-up — both teams evenly matched"),
    ]

    fills = [GREEN_FILL, GREEN_FILL, YELLOW_FILL, YELLOW_FILL,
             ORANGE_FILL, ORANGE_FILL, PURPLE_FILL, BLUE_FILL]

    for ri, (row_data, f) in enumerate(zip(data, fills), 3):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            dc(cell, f if ci in (1,3,5) else None)
        ws.row_dimensions[ri].height = 20

    for i, w in enumerate([12,16,10,14,28,40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── SHEET 6: HOW TO USE ───────────────────────────────────────────────────────
def build_how_to_sheet(wb):
    ws = wb.create_sheet("ℹ️ How To Use")
    ws.sheet_view.showGridLines = False
    title_row(ws, "ℹ️  HOW TO USE THIS TOOL", 2)
    lines = [
        ("DAILY WORKFLOW:", True),
        ("1.  Run:  python3 cbb_betting_model.py  in Terminal each morning", False),
        ("2.  Open the Excel file created in your MLB Betting folder", False),
        ("3.  📊 Today's Matchups — main stat-based recommendations", False),
        ("4.  🎯 Parlay Suggestions — best combo bets", False),
        ("5.  🎲 Blind Underdog Bets — 2u on every dog getting 4+ pts (fun/experimental)", False),
        ("6.  📋 Team Cheat Sheet — full team ratings reference", False),
        ("7.  🏀 Seed Matchup History — use during March Madness for context", False),
        ("", False),
        ("STAT MODEL BETS (confident money):", True),
        ("•  🟢 STRONG FAV (Gap 30+) → Bet alt spread -4.5 to -6.5, include in parlays", False),
        ("•  🟡 LEAN FAV (Gap 18–29) → Alt -4.5 only, half size", False),
        ("•  ⭐ PRIME UPSET → Bet underdog ML, especially 5v12, 6v11, 7v10 in tournament", False),
        ("•  ⚪ TOSS-UP → Pass unless you have a strong lean", False),
        ("", False),
        ("BLIND UNDERDOG STRATEGY:", True),
        ("•  Every game where dog is getting 4+ pts → 2 units on dog ML", False),
        ("•  This is intentionally dumb and fun — track it in the running P&L table", False),
        ("•  Best case: you find the 5v12 and 6v11 upsets early and profit big", False),
        ("•  Expected outcome over a full season: slight negative, but fun ROI in March", False),
        ("", False),
        ("MARCH MADNESS SPECIFIC TIPS:", True),
        ("•  12 seeds vs 5 seeds: always bet the 12 blind at 2u — hits 35% of the time", False),
        ("•  11 seeds vs 6 seeds: same logic — historically the best upset matchup", False),
        ("•  Experience matters enormously — filter for senior-heavy teams in tight games", False),
        ("•  Efficiency margin gap of 10+ with spread under 6 = massive value on favorite", False),
        ("", False),
        ("KEEPING STATS FRESH:", True),
        ("•  KenPom efficiency data: kenpom.com (~$25/yr subscription, best investment)", False),
        ("•  Free alternative: barttorvik.com (T-Rank) — nearly as good, totally free", False),
        ("•  Update TEAM_STATS dict in the script weekly during the season", False),
        ("•  Odds auto-update every time you run the script via the API", False),
    ]
    for ri, (text, bold) in enumerate(lines, 3):
        c = ws.cell(row=ri, column=1, value=text)
        c.font = Font(bold=bold, size=11)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=2)
        ws.row_dimensions[ri].height = 18 if not bold else 22
    ws.column_dimensions["A"].width = 80


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "="*55)
    print("  🏀  CBB DAILY BETTING MODEL")
    print(f"  {date.today().strftime('%A, %B %d, %Y')}")
    print("="*55 + "\n")

    odds_data = fetch_cbb_odds()

    if not odds_data:
        print("\n⚠️  No games found today (off-day or API issue).")
        print("   Building cheat sheets only...\n")
        games = blind_dogs = parlays = []
    else:
        games      = parse_games(odds_data)
        blind_dogs = get_blind_dogs(games)
        parlays    = suggest_parlays(games)

    wb = Workbook()
    build_matchup_sheet(wb, games)
    build_parlay_sheet(wb, parlays)
    build_blind_dog_sheet(wb, blind_dogs)
    build_cheat_sheet(wb)
    build_seed_history_sheet(wb)
    build_how_to_sheet(wb)
    wb.save(OUTPUT_FILE)

    print(f"\n✅  Done! File saved: {OUTPUT_FILE}")
    print(f"   Games today:        {len(games)}")
    strong = [g for g in games if "STRONG" in g["Rating"]]
    upsets = [g for g in games if "PRIME"  in g["Upset Flag"]]
    print(f"   🟢 Strong fav bets: {len(strong)}")
    print(f"   ⭐ Prime upsets:    {len(upsets)}")
    print(f"   🎲 Blind dogs:      {len(blind_dogs)}")
    print(f"   🎯 Parlay combos:   {len(parlays)}")

    if strong:
        print("\n  TOP STAT PICKS:")
        for g in strong[:3]:
            print(f"  ⭐ {g['Favorite']} vs {g['Underdog']}  |  Gap: {g['Gap']}  |  {g['Alt Spread']}")
    if upsets:
        print("\n  UPSET ALERTS:")
        for g in upsets[:3]:
            print(f"  🎲 {g['Underdog']} (+{g['Spread'].split()[-1] if g['Spread'] != 'N/A' else '?'})  vs {g['Favorite']}  |  Hist. rate: {g['Upset Rate']}")
    print()


if __name__ == "__main__":
    main()
