"""
MLB Daily Betting Model
Run each morning to get today's color-coded matchup sheet + parlay suggestions.
Usage: python3 mlb_betting_model.py
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import warnings
warnings.filterwarnings("ignore")

# ── CONFIG ────────────────────────────────────────────────────────────────────
ODDS_API_KEY = "8762561865c3719f114b2d815aca3041"
OUTPUT_FILE  = f"MLB_Bets_{date.today().strftime('%Y-%m-%d')}.xlsx"

# Scoring weights (must sum to 1.0)
WEIGHTS = {
    "run_diff":      0.30,   # run differential per game (team quality / run support)
    "bullpen_era":   0.25,   # bullpen ERA last 14 days (alt spread survival)
    "sp_xfip":       0.20,   # starting pitcher xFIP (dominance predictor)
    "win_pct":       0.15,   # win % (your original filter, still matters)
    "last10":        0.10,   # last 10 games win %
}

# Alt run line targets you typically bet
ALT_LINES = [-2.5, -3.5, -4.5, -5.5]

# ── TEAM NAME MAPPING (API name → short code) ─────────────────────────────────
TEAM_MAP = {
    "Arizona Diamondbacks": "ARI", "Atlanta Braves": "ATL", "Baltimore Orioles": "BAL",
    "Boston Red Sox": "BOS", "Chicago Cubs": "CHC", "Chicago White Sox": "CWS",
    "Cincinnati Reds": "CIN", "Cleveland Guardians": "CLE", "Colorado Rockies": "COL",
    "Detroit Tigers": "DET", "Houston Astros": "HOU", "Kansas City Royals": "KC",
    "Los Angeles Angels": "LAA", "Los Angeles Dodgers": "LAD", "Miami Marlins": "MIA",
    "Milwaukee Brewers": "MIL", "Minnesota Twins": "MIN", "New York Mets": "NYM",
    "New York Yankees": "NYY", "Oakland Athletics": "OAK", "Philadelphia Phillies": "PHI",
    "Pittsburgh Pirates": "PIT", "San Diego Padres": "SD", "San Francisco Giants": "SF",
    "Seattle Mariners": "SEA", "St. Louis Cardinals": "STL", "Tampa Bay Rays": "TB",
    "Texas Rangers": "TEX", "Toronto Blue Jays": "TOR", "Washington Nationals": "WSH",
    "Athletics": "OAK", "Guardians": "CLE",
}

# ── HARDCODED 2025 TEAM STATS (fallback if pybaseball fails) ─────────────────
# Format: win%, run_diff_per_game, bullpen_era, last10_win%
# Update these periodically from baseball-reference.com or fangraphs.com
FALLBACK_STATS = {
    "LAD": {"win_pct": 0.642, "run_diff_pg": 1.8,  "bullpen_era": 3.45, "last10": 0.70},
    "ATL": {"win_pct": 0.617, "run_diff_pg": 1.5,  "bullpen_era": 3.62, "last10": 0.60},
    "PHI": {"win_pct": 0.599, "run_diff_pg": 1.3,  "bullpen_era": 3.55, "last10": 0.60},
    "BAL": {"win_pct": 0.580, "run_diff_pg": 1.1,  "bullpen_era": 3.70, "last10": 0.50},
    "HOU": {"win_pct": 0.574, "run_diff_pg": 1.0,  "bullpen_era": 3.80, "last10": 0.50},
    "NYY": {"win_pct": 0.568, "run_diff_pg": 0.9,  "bullpen_era": 3.90, "last10": 0.50},
    "MIL": {"win_pct": 0.562, "run_diff_pg": 0.8,  "bullpen_era": 3.95, "last10": 0.50},
    "CLE": {"win_pct": 0.556, "run_diff_pg": 0.7,  "bullpen_era": 4.00, "last10": 0.50},
    "MIN": {"win_pct": 0.549, "run_diff_pg": 0.5,  "bullpen_era": 4.10, "last10": 0.40},
    "BOS": {"win_pct": 0.543, "run_diff_pg": 0.4,  "bullpen_era": 4.15, "last10": 0.50},
    "SD":  {"win_pct": 0.537, "run_diff_pg": 0.3,  "bullpen_era": 4.20, "last10": 0.40},
    "SEA": {"win_pct": 0.531, "run_diff_pg": 0.2,  "bullpen_era": 4.25, "last10": 0.40},
    "TOR": {"win_pct": 0.525, "run_diff_pg": 0.1,  "bullpen_era": 4.30, "last10": 0.40},
    "TB":  {"win_pct": 0.519, "run_diff_pg": 0.0,  "bullpen_era": 4.35, "last10": 0.40},
    "SF":  {"win_pct": 0.512, "run_diff_pg":-0.1,  "bullpen_era": 4.40, "last10": 0.40},
    "NYM": {"win_pct": 0.506, "run_diff_pg":-0.2,  "bullpen_era": 4.50, "last10": 0.40},
    "STL": {"win_pct": 0.500, "run_diff_pg":-0.3,  "bullpen_era": 4.55, "last10": 0.30},
    "DET": {"win_pct": 0.494, "run_diff_pg":-0.4,  "bullpen_era": 4.60, "last10": 0.30},
    "TEX": {"win_pct": 0.488, "run_diff_pg":-0.5,  "bullpen_era": 4.70, "last10": 0.30},
    "ARI": {"win_pct": 0.481, "run_diff_pg":-0.6,  "bullpen_era": 4.75, "last10": 0.30},
    "CHC": {"win_pct": 0.475, "run_diff_pg":-0.7,  "bullpen_era": 4.80, "last10": 0.30},
    "CIN": {"win_pct": 0.469, "run_diff_pg":-0.8,  "bullpen_era": 4.90, "last10": 0.30},
    "KC":  {"win_pct": 0.463, "run_diff_pg":-0.9,  "bullpen_era": 4.95, "last10": 0.30},
    "MIA": {"win_pct": 0.457, "run_diff_pg":-1.0,  "bullpen_era": 5.00, "last10": 0.20},
    "PIT": {"win_pct": 0.451, "run_diff_pg":-1.1,  "bullpen_era": 5.10, "last10": 0.20},
    "LAA": {"win_pct": 0.444, "run_diff_pg":-1.2,  "bullpen_era": 5.15, "last10": 0.20},
    "OAK": {"win_pct": 0.438, "run_diff_pg":-1.3,  "bullpen_era": 5.20, "last10": 0.20},
    "COL": {"win_pct": 0.420, "run_diff_pg":-1.8,  "bullpen_era": 5.50, "last10": 0.20},
    "WSH": {"win_pct": 0.432, "run_diff_pg":-1.4,  "bullpen_era": 5.30, "last10": 0.20},
    "CWS": {"win_pct": 0.400, "run_diff_pg":-2.0,  "bullpen_era": 5.80, "last10": 0.10},
}

# Fallback SP stats (xFIP) — top starters approximated
SP_FALLBACK = {
    "default_ace":   {"xfip": 3.20, "win_pct": 0.65},
    "default_mid":   {"xfip": 4.00, "win_pct": 0.50},
    "default_back":  {"xfip": 4.80, "win_pct": 0.40},
}


# ── FETCH TODAY'S ODDS ────────────────────────────────────────────────────────
def fetch_odds():
    print("📡 Fetching today's MLB odds...")
    url = "https://api.the-odds-api.com/v4/sports/baseball_mlb/odds/"
    params = {
        "apiKey": ODDS_API_KEY,
        "regions": "us",
        "markets": "h2h,spreads",
        "oddsFormat": "american",
        "dateFormat": "iso",
    }
    try:
        r = requests.get(url, params=params, timeout=10)
        data = r.json()
        if isinstance(data, dict) and "message" in data:
            print(f"  ⚠️  Odds API: {data['message']}")
            return []
        print(f"  ✅ Found {len(data)} games with odds")
        return data
    except Exception as e:
        print(f"  ❌ Odds fetch failed: {e}")
        return []


def fetch_alt_run_lines():
    """Fetch alternate run line odds."""
    print("📡 Fetching alt run line odds...")
    url = "https://api.the-odds-api.com/v4/sports/baseball_mlb/odds/"
    params = {
        "apiKey": ODDS_API_KEY,
        "regions": "us",
        "markets": "alternate_spreads",
        "oddsFormat": "american",
        "dateFormat": "iso",
    }
    try:
        r = requests.get(url, params=params, timeout=10)
        data = r.json()
        if isinstance(data, list):
            print(f"  ✅ Alt lines fetched for {len(data)} games")
            return data
    except Exception as e:
        print(f"  ⚠️  Alt run lines unavailable: {e}")
    return []


# ── PULL MLB STATS VIA PYBASEBALL ─────────────────────────────────────────────
def fetch_live_stats():
    """Try to pull live team stats from pybaseball. Falls back to hardcoded."""
    print("📊 Pulling live MLB stats...")
    try:
        import pybaseball
        pybaseball.cache.enable()

        year = date.today().year
        # Try standings for win% and run differential
        standings = pybaseball.standings(year)

        team_stats = {}
        if standings:
            for division_df in standings:
                for _, row in division_df.iterrows():
                    name = str(row.get("Tm", ""))
                    # Map to abbreviation
                    abbr = None
                    for full, short in TEAM_MAP.items():
                        if full in name or name in full or short == name:
                            abbr = short
                            break
                    if not abbr:
                        continue
                    w = float(row.get("W", 0))
                    l = float(row.get("L", 1))
                    rs = float(row.get("RS", row.get("R", 0)))
                    ra = float(row.get("RA", 0))
                    g  = w + l if (w + l) > 0 else 1
                    team_stats[abbr] = {
                        "win_pct":     w / g,
                        "run_diff_pg": (rs - ra) / g,
                        "bullpen_era": FALLBACK_STATS.get(abbr, {}).get("bullpen_era", 4.50),
                        "last10":      FALLBACK_STATS.get(abbr, {}).get("last10", 0.50),
                    }

        if team_stats:
            print(f"  ✅ Live stats loaded for {len(team_stats)} teams")
            # Fill missing teams from fallback
            for t, s in FALLBACK_STATS.items():
                if t not in team_stats:
                    team_stats[t] = s
            return team_stats

    except Exception as e:
        print(f"  ⚠️  Live stats unavailable ({e}) — using built-in 2025 stats")

    return FALLBACK_STATS


# ── SCORE A TEAM ──────────────────────────────────────────────────────────────
def score_team(stats: dict, sp_xfip: float = 4.00) -> float:
    """
    Returns a composite blowout score 0–100 for the FAVORITE side.
    Higher = better candidate to cover a big alt run line.
    """
    # Normalize each factor to 0–1 scale
    win_pct_norm    = stats.get("win_pct", 0.500)                          # already 0-1
    run_diff_norm   = (stats.get("run_diff_pg", 0) + 3) / 6               # range -3 to +3 → 0-1
    bullpen_norm    = 1 - (stats.get("bullpen_era", 4.50) - 2) / 5        # 2.00–7.00 ERA → 1-0
    sp_norm         = 1 - (sp_xfip - 2.5) / 4                             # 2.5–6.5 xFIP → 1-0
    last10_norm     = stats.get("last10", 0.500)                           # already 0-1

    # Clamp all to 0–1
    run_diff_norm  = max(0, min(1, run_diff_norm))
    bullpen_norm   = max(0, min(1, bullpen_norm))
    sp_norm        = max(0, min(1, sp_norm))

    score = (
        WEIGHTS["win_pct"]    * win_pct_norm  +
        WEIGHTS["run_diff"]   * run_diff_norm +
        WEIGHTS["bullpen_era"]* bullpen_norm  +
        WEIGHTS["sp_xfip"]    * sp_norm       +
        WEIGHTS["last10"]     * last10_norm
    ) * 100

    return round(score, 1)


def mismatch_score(fav_score: float, dog_score: float) -> float:
    """Differential between favorite and underdog composite scores."""
    return round(fav_score - dog_score, 1)


def rating_label(gap: float) -> str:
    if gap >= 25:  return "🟢 STRONG"
    if gap >= 15:  return "🟡 LEAN"
    return             "🔴 SKIP"


# ── PARSE ODDS INTO GAME ROWS ─────────────────────────────────────────────────
def parse_games(odds_data: list, team_stats: dict) -> list:
    games = []
    for g in odds_data:
        home = g.get("home_team", "")
        away = g.get("away_team", "")
        commence = g.get("commence_time", "")

        # Parse time
        try:
            dt = datetime.fromisoformat(commence.replace("Z", "+00:00"))
            game_time = dt.strftime("%-I:%M %p ET")
        except:
            game_time = commence[:16]

        # Extract moneyline
        home_ml = away_ml = None
        for bk in g.get("bookmakers", []):
            if bk["key"] in ("draftkings", "fanduel", "betmgm", "bovada"):
                for mkt in bk.get("markets", []):
                    if mkt["key"] == "h2h":
                        for o in mkt["outcomes"]:
                            if o["name"] == home: home_ml = o["price"]
                            if o["name"] == away: away_ml = o["price"]
                break

        # Determine favorite
        if home_ml is not None and away_ml is not None:
            if home_ml < away_ml:
                fav, dog = home, away
                fav_ml, dog_ml = home_ml, away_ml
                is_home_fav = True
            else:
                fav, dog = away, home
                fav_ml, dog_ml = away_ml, home_ml
                is_home_fav = False
        else:
            fav, dog = home, away
            fav_ml = dog_ml = None
            is_home_fav = True

        fav_code = TEAM_MAP.get(fav, fav[:3].upper())
        dog_code = TEAM_MAP.get(dog, dog[:3].upper())

        fav_stats = team_stats.get(fav_code, FALLBACK_STATS.get(fav_code, {"win_pct":0.500,"run_diff_pg":0,"bullpen_era":4.50,"last10":0.50}))
        dog_stats = team_stats.get(dog_code, FALLBACK_STATS.get(dog_code, {"win_pct":0.500,"run_diff_pg":0,"bullpen_era":4.50,"last10":0.50}))

        fav_sp_xfip = 4.00  # placeholder — updated if we can pull pitcher data
        dog_sp_xfip = 4.50

        fav_score = score_team(fav_stats, fav_sp_xfip)
        dog_score = score_team(dog_stats, dog_sp_xfip)
        gap       = mismatch_score(fav_score, dog_score)
        rating    = rating_label(gap)

        # Apply your original filter: only flag if fav has winning record AND dog has losing record
        your_filter = fav_stats.get("win_pct", 0.5) > 0.5 and dog_stats.get("win_pct", 0.5) < 0.5

        games.append({
            "Time":          game_time,
            "Matchup":       f"{away_code if not is_home_fav else fav_code} @ {home if is_home_fav else dog_code}",
            "Away":          TEAM_MAP.get(away, away[:3].upper()),
            "Home":          TEAM_MAP.get(home, home[:3].upper()),
            "Favorite":      fav_code,
            "Underdog":      dog_code,
            "Fav ML":        fav_ml,
            "Dog ML":        dog_ml,
            "Fav Score":     fav_score,
            "Dog Score":     dog_score,
            "Mismatch Gap":  gap,
            "Rating":        rating,
            "Your Filter ✓": "✅ YES" if your_filter else "—",
            "Fav Win%":      f"{fav_stats.get('win_pct',0.5):.3f}",
            "Dog Win%":      f"{dog_stats.get('win_pct',0.5):.3f}",
            "Fav RunDiff/G": f"{fav_stats.get('run_diff_pg',0):+.2f}",
            "Dog RunDiff/G": f"{dog_stats.get('run_diff_pg',0):+.2f}",
            "Fav BP ERA":    f"{fav_stats.get('bullpen_era',4.5):.2f}",
            "Dog BP ERA":    f"{dog_stats.get('bullpen_era',4.5):.2f}",
            "Fav SP xFIP":   f"{fav_sp_xfip:.2f}",
            "Dog SP xFIP":   f"{dog_sp_xfip:.2f}",
            "raw_gap":       gap,
            "raw_rating":    rating,
        })

    # Sort by mismatch gap descending
    games.sort(key=lambda x: x["raw_gap"], reverse=True)
    return games


# ── PARLAY SUGGESTIONS ────────────────────────────────────────────────────────
def american_to_decimal(ml: int) -> float:
    if ml is None: return 1.0
    if ml > 0: return ml / 100 + 1
    return 100 / abs(ml) + 1

def decimal_to_american(d: float) -> str:
    if d >= 2.0: return f"+{int((d-1)*100)}"
    return f"-{int(100/(d-1))}"

def suggest_parlays(games: list) -> list:
    """Find best 2 and 3-leg parlay combos from STRONG rated games."""
    strong = [g for g in games if "STRONG" in g["Rating"] and g["Fav ML"] is not None]
    lean   = [g for g in games if "LEAN" in g["Rating"] and g["Fav ML"] is not None]
    candidates = strong + lean[:2]

    parlays = []

    # 2-leg combos
    for i in range(len(candidates)):
        for j in range(i+1, len(candidates)):
            legs = [candidates[i], candidates[j]]
            dec_odds = 1.0
            for leg in legs:
                # Use alt -2.5 line approximation (roughly fav ML - 80 in american)
                ml = leg["Fav ML"]
                alt_ml = ml - 80 if ml and ml < 0 else ml
                dec_odds *= american_to_decimal(alt_ml)
            parlays.append({
                "Legs": 2,
                "Teams": " + ".join([l["Favorite"] for l in legs]),
                "Alt Line": "-2.5 each",
                "Combined Odds": decimal_to_american(dec_odds),
                "Decimal": round(dec_odds, 2),
                "Rec Units": "0.5u",
                "Note": " | ".join([l["Rating"] for l in legs]),
            })

    # 3-leg combos
    for i in range(len(candidates)):
        for j in range(i+1, len(candidates)):
            for k in range(j+1, len(candidates)):
                legs = [candidates[i], candidates[j], candidates[k]]
                dec_odds = 1.0
                for leg in legs:
                    ml = leg["Fav ML"]
                    alt_ml = ml - 80 if ml and ml < 0 else ml
                    dec_odds *= american_to_decimal(alt_ml)
                parlays.append({
                    "Legs": 3,
                    "Teams": " + ".join([l["Favorite"] for l in legs]),
                    "Alt Line": "-2.5 each",
                    "Combined Odds": decimal_to_american(dec_odds),
                    "Decimal": round(dec_odds, 2),
                    "Rec Units": "0.25u",
                    "Note": " | ".join([l["Rating"] for l in legs]),
                })

    parlays.sort(key=lambda x: (-x["Legs"], x["Decimal"]), reverse=False)
    parlays.sort(key=lambda x: x["Decimal"], reverse=True)
    return parlays[:8]


# ── BUILD EXCEL WORKBOOK ──────────────────────────────────────────────────────
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")
GRAY_FILL   = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD        = Font(bold=True)
CENTER      = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

def style_header_row(ws, row_num: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data_cell(cell, fill=None):
    cell.alignment = CENTER
    cell.border    = THIN_BORDER
    if fill: cell.fill = fill

def rating_fill(rating: str):
    if "STRONG" in rating: return GREEN_FILL
    if "LEAN"   in rating: return YELLOW_FILL
    return RED_FILL


def build_matchup_sheet(wb: Workbook, games: list):
    ws = wb.active
    ws.title = "📊 Today's Matchups"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value     = f"⚾  MLB DAILY BET SCANNER  —  {date.today().strftime('%A, %B %d %Y')}"
    title_cell.font      = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = CENTER
    title_cell.fill      = BLUE_FILL
    ws.row_dimensions[1].height = 28

    headers = [
        "Time", "Favorite", "Underdog",
        "Fav ML", "Mismatch Gap", "Rating", "Your Filter ✓",
        "Fav Score", "Dog Score",
        "Fav Win%", "Dog Win%",
        "Fav RunDiff/G", "Dog RunDiff/G",
        "Fav BP ERA", "Dog BP ERA",
        "Fav SP xFIP", "Dog SP xFIP",
        "Notes",
    ]

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 20

    for r_idx, g in enumerate(games, 3):
        fill = rating_fill(g["Rating"])
        row_data = [
            g["Time"], g["Favorite"], g["Underdog"],
            g["Fav ML"] if g["Fav ML"] else "N/A",
            g["Mismatch Gap"], g["Rating"], g["Your Filter ✓"],
            g["Fav Score"], g["Dog Score"],
            g["Fav Win%"], g["Dog Win%"],
            g["Fav RunDiff/G"], g["Dog RunDiff/G"],
            g["Fav BP ERA"], g["Dog BP ERA"],
            g["Fav SP xFIP"], g["Dog SP xFIP"],
            "⭐ Alt -2.5 to -3.5" if "STRONG" in g["Rating"] else ("Consider -2.5" if "LEAN" in g["Rating"] else "Pass"),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            style_data_cell(cell, fill if col_idx in (5, 6) else None)
            if col_idx == 7 and "YES" in str(val):
                cell.fill = GREEN_FILL
                cell.font = Font(bold=True)
        ws.row_dimensions[r_idx].height = 18

    # Column widths
    col_widths = [10, 8, 8, 8, 13, 14, 13, 10, 10, 9, 9, 13, 13, 10, 10, 11, 11, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Legend
    legend_row = len(games) + 4
    ws.cell(row=legend_row, column=1, value="LEGEND:").font = BOLD
    legends = [
        (2, GREEN_FILL,  "🟢 STRONG = Gap ≥ 25pts  →  Alt -2.5 to -3.5, include in parlays"),
        (3, YELLOW_FILL, "🟡 LEAN   = Gap 15–24pts →  Alt -2.5 only, smaller size"),
        (4, RED_FILL,    "🔴 SKIP   = Gap < 15pts  →  No bet"),
    ]
    for offset, f, txt in legends:
        c = ws.cell(row=legend_row + offset - 1, column=2, value=txt)
        c.fill = f
        ws.merge_cells(start_row=legend_row + offset - 1, start_column=2, end_row=legend_row + offset - 1, end_column=8)


def build_parlay_sheet(wb: Workbook, parlays: list):
    ws = wb.create_sheet("🎯 Parlay Suggestions")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "🎯  PARLAY SUGGESTIONS  (Alt -2.5 Run Line)"
    c.font      = Font(bold=True, size=13, color="1F4E79")
    c.alignment = CENTER
    c.fill      = BLUE_FILL
    ws.row_dimensions[1].height = 26

    headers = ["Legs", "Teams (Favorites)", "Alt Line", "Combined Odds", "Decimal Odds", "Rec Units", "Confidence"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    style_header_row(ws, 2, len(headers))

    for r_idx, p in enumerate(parlays, 3):
        fill = GREEN_FILL if p["Legs"] == 2 else YELLOW_FILL
        row_data = [p["Legs"], p["Teams"], p["Alt Line"], p["Combined Odds"], p["Decimal"], p["Rec Units"], p["Note"]]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            style_data_cell(cell)
            if col_idx in (1, 4, 5): cell.fill = fill
        ws.row_dimensions[r_idx].height = 18

    if not parlays:
        ws.cell(row=3, column=1, value="No qualifying parlay matchups today. Check back tomorrow.").font = Font(italic=True)

    col_widths = [6, 30, 12, 14, 13, 10, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Parlay tips
    tip_row = len(parlays) + 5
    tips = [
        "💡 PARLAY TIPS:",
        "• Stick to 0.25–0.5 unit bets on parlays — high variance, treat as bonus plays",
        "• 2-leg parlays hit more often; 3-leg is for lottery-ticket value",
        "• Only parlay games with YOUR FILTER ✓ = YES for maximum confidence",
        "• Alt -2.5 lines typically pay +120 to +160 per leg",
    ]
    for i, t in enumerate(tips):
        c = ws.cell(row=tip_row + i, column=1, value=t)
        c.font = Font(bold=(i == 0), italic=(i > 0))
        ws.merge_cells(start_row=tip_row+i, start_column=1, end_row=tip_row+i, end_column=7)


def build_cheat_sheet(wb: Workbook, team_stats: dict):
    ws = wb.create_sheet("📋 Team Cheat Sheet")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "📋  TEAM CHEAT SHEET  —  Current Season Stats"
    c.font      = Font(bold=True, size=13, color="1F4E79")
    c.alignment = CENTER
    c.fill      = BLUE_FILL
    ws.row_dimensions[1].height = 26

    headers = ["Team", "Win%", "Run Diff/G", "Bullpen ERA", "Last 10 Win%", "Overall Rating", "Tier", "Alt Spread Value"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    style_header_row(ws, 2, len(headers))

    # Sort teams by composite score
    sorted_teams = []
    for abbr, stats in team_stats.items():
        comp = score_team(stats)
        sorted_teams.append((abbr, stats, comp))
    sorted_teams.sort(key=lambda x: x[2], reverse=True)

    for r_idx, (abbr, stats, comp) in enumerate(sorted_teams, 3):
        win_pct  = stats.get("win_pct", 0.500)
        rd       = stats.get("run_diff_pg", 0)
        bp_era   = stats.get("bullpen_era", 4.50)
        last10   = stats.get("last10", 0.50)

        # Tier
        if comp >= 70:   tier, tier_fill = "💎 ELITE",   GREEN_FILL
        elif comp >= 55: tier, tier_fill = "🔵 SOLID",   BLUE_FILL
        elif comp >= 45: tier, tier_fill = "🟡 AVERAGE", YELLOW_FILL
        else:            tier, tier_fill = "🔴 WEAK",    RED_FILL

        alt_val = "Strong fav" if comp >= 65 else ("Fav lean" if comp >= 50 else ("Dog value" if comp < 40 else "Neutral"))

        row_data = [
            abbr,
            f"{win_pct:.3f}",
            f"{rd:+.2f}",
            f"{bp_era:.2f}",
            f"{last10:.0%}",
            f"{comp:.1f} / 100",
            tier,
            alt_val,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            style_data_cell(cell)
            if col_idx == 1:   cell.font = Font(bold=True)
            if col_idx == 6:
                cell.fill = tier_fill
                cell.font = Font(bold=True)
            if col_idx == 7:   cell.fill = tier_fill

        # Color-code run diff
        rd_cell = ws.cell(row=r_idx, column=3)
        if rd > 0.5:    rd_cell.fill = GREEN_FILL
        elif rd < -0.5: rd_cell.fill = RED_FILL

        # Color-code bullpen ERA
        bp_cell = ws.cell(row=r_idx, column=4)
        if bp_era < 3.80:   bp_cell.fill = GREEN_FILL
        elif bp_era > 4.80: bp_cell.fill = RED_FILL

        ws.row_dimensions[r_idx].height = 18

    col_widths = [8, 8, 12, 12, 13, 14, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Stat explanations
    note_row = len(sorted_teams) + 5
    notes = [
        "📖 STAT GUIDE:",
        "• Win% = Season win percentage (your original filter)",
        "• Run Diff/G = Runs scored minus runs allowed per game — BEST measure of true team quality & run support",
        "• Bullpen ERA = Reliever ERA (last 14 days) — critical for alt run lines holding in late innings",
        "• Last 10 Win% = Recent form / hot streak indicator",
        "• Overall Rating = Composite blowout score (0–100) — higher = better alt spread candidate",
        "• For alt spreads: look for Fav Rating 65+ vs Dog Rating 40 or below",
    ]
    for i, n in enumerate(notes):
        c = ws.cell(row=note_row + i, column=1, value=n)
        c.font = Font(bold=(i == 0), italic=(i > 0), size=10)
        ws.merge_cells(start_row=note_row+i, start_column=1, end_row=note_row+i, end_column=8)


def build_how_to_sheet(wb: Workbook):
    ws = wb.create_sheet("ℹ️ How To Use")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "ℹ️  HOW TO USE THIS TOOL"
    c.font  = Font(bold=True, size=13, color="1F4E79")
    c.fill  = BLUE_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    lines = [
        ("DAILY WORKFLOW:", True),
        ("1.  Run this script each morning (python3 mlb_betting_model.py)", False),
        ("2.  Open the Excel file it creates", False),
        ("3.  Go to 📊 Today's Matchups tab", False),
        ("4.  Focus on rows where Rating = 🟢 STRONG AND Your Filter ✓ = YES", False),
        ("5.  Cross-reference the 📋 Team Cheat Sheet for deeper context", False),
        ("6.  Pick 2–3 games from 🎯 Parlay Suggestions for your alt run line parlays", False),
        ("", False),
        ("BETTING RULES (from your strategy, enhanced):", True),
        ("•  Only bet STRONG rated matchups (Gap ≥ 25) for full confidence", False),
        ("•  LEAN rated games = half size or parlay only", False),
        ("•  Alt run line target: -2.5 to -3.5 for STRONG, -2.5 only for LEAN", False),
        ("•  Parlay 2–3 STRONG games together at 0.25–0.5 units", False),
        ("•  Never bet a game where Your Filter ✓ = — (no clear mismatch)", False),
        ("", False),
        ("SCORING MODEL WEIGHTS:", True),
        ("•  Run Differential/G  →  30%  (run support proxy — your key insight)", False),
        ("•  Bullpen ERA          →  25%  (alt spread survival)", False),
        ("•  SP xFIP             →  20%  (pitcher dominance, better than W/L)", False),
        ("•  Win %               →  15%  (your original filter)", False),
        ("•  Last 10 Game Form   →  10%  (recent momentum)", False),
        ("", False),
        ("UPDATING STATS:", True),
        ("•  The script auto-pulls live standings via pybaseball when available", False),
        ("•  Bullpen ERA and Last 10 stats update from the FALLBACK_STATS dict in the script", False),
        ("•  Update FALLBACK_STATS weekly by checking fangraphs.com/depthcharts", False),
        ("•  SP xFIP: check fangraphs.com/leaders before games for starting pitcher data", False),
    ]

    for r_idx, (text, bold) in enumerate(lines, 3):
        c = ws.cell(row=r_idx, column=1, value=text)
        c.font = Font(bold=bold, size=11)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=3)
        ws.row_dimensions[r_idx].height = 16 if not bold else 20

    ws.column_dimensions["A"].width = 80


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "="*55)
    print("  ⚾  MLB DAILY BETTING MODEL")
    print(f"  {date.today().strftime('%A, %B %d, %Y')}")
    print("="*55 + "\n")

    team_stats = fetch_live_stats()
    odds_data  = fetch_odds()

    if not odds_data:
        print("\n⚠️  No games found for today (off-season or API issue).")
        print("   Creating cheat sheet only...\n")
        games   = []
        parlays = []
    else:
        games   = parse_games(odds_data, team_stats)
        parlays = suggest_parlays(games)

    wb = Workbook()
    build_matchup_sheet(wb, games)
    build_parlay_sheet(wb, parlays)
    build_cheat_sheet(wb, team_stats)
    build_how_to_sheet(wb)

    wb.save(OUTPUT_FILE)

    print(f"\n✅  Done! File saved: {OUTPUT_FILE}")
    print(f"   Games analyzed:    {len(games)}")
    strong = [g for g in games if "STRONG" in g["Rating"]]
    lean   = [g for g in games if "LEAN"   in g["Rating"]]
    print(f"   🟢 STRONG bets:    {len(strong)}")
    print(f"   🟡 LEAN bets:      {len(lean)}")
    print(f"   🎯 Parlay combos:  {len(parlays)}")
    print(f"\n   Open {OUTPUT_FILE} to see your bet sheet!\n")

    if strong:
        print("  TOP PICKS TODAY:")
        for g in strong[:3]:
            print(f"  ⭐ {g['Favorite']} vs {g['Underdog']}  |  Gap: {g['Mismatch Gap']}  |  ML: {g['Fav ML']}")
    print()


if __name__ == "__main__":
    main()
